from openpyxl import *
from textblob import *
from statistics import *
from heapq import *
import os

# Defining two functions called six_table_creatorReviews & six_table_creatorSents. These functions take several arguments and displays them into a table tbat label the descriptive statistics of lists

def six_table_creatorReviews (array1, array2, a1, s1, n1, m1, l1, a2, s2, n2, m2, l2, a3, s3, n3, m3, l3, a4, s4, n4, m4, l4, a5, s5, n5, m5, l5, a6, s6, n6, m6, l6):      
    print("Descriptive Statistics Table for the top and bottom 3 average reviews")
    print("           Mean  Standard Deviation   Sample Size   Max   Min")
    print(array1[0],"{:>2.2f} {:>19.2f} {:>13.0f} {:>6.2f} {:>5.2f}" .format(a1,s1,n1,m1,l1))
    print(array1[1],"{:>7.2f} {:>19.2f} {:>13.0f} {:>6.2f} {:>5.2f}" .format(a2,s2,n2,n2,l2))
    print(array1[2],"{:>2.2f} {:>19.2f} {:>13.0f} {:>6.2f} {:>5.2f}" .format(a3,s3,n3,n3,l3))
    print(array2[0],"{:>5.2f} {:>19.2f} {:>13.0f} {:>6.2f} {:>5.2f}" .format(a4,s4,n4,n4,l4))
    print(array2[1],"{:>7.2f} {:>19.2f} {:>13.0f} {:>6.2f} {:>5.2f}" .format(a5,s5,n5,n5,l5))
    print(array2[2],"{:>7.2f} {:>19.2f} {:>13.0f} {:>6.2f} {:>5.2f}" .format(a6,s6,n6,n6,l6))

def six_table_creatorSents (array1, array2, a1, s1, n1, m1, l1, a2, s2, n2, m2, l2, a3, s3, n3, m3, l3, a4, s4, n4, m4, l4, a5, s5, n5, m5, l5, a6, s6, n6, m6, l6):
    print("Descriptive Statistics Table for the top and bottom 3 average sentiments")
    print("           Mean  Standard Deviation   Sample Size   Max   Min")
    print(array1[0],"{:>2.2f} {:>19.2f} {:>13.0f} {:>6.2f} {:>5.2f}" .format(a1,s1,n1,m1,l1))
    print(array1[1],"{:>7.2f} {:>19.2f} {:>13.0f} {:>6.2f} {:>5.2f}" .format(a2,s2,n2,n2,l2))
    print(array1[2],"{:>2.2f} {:>19.2f} {:>13.0f} {:>6.2f} {:>5.2f}" .format(a3,s3,n3,n3,l3))
    print(array2[0],"{:>5.2f} {:>19.2f} {:>13.0f} {:>6.2f} {:>5.2f}" .format(a4,s4,n4,n4,l4))
    print(array2[1],"{:>7.2f} {:>19.2f} {:>13.0f} {:>6.2f} {:>5.2f}" .format(a5,s5,n5,n5,l5))
    print(array2[2],"{:>7.2f} {:>19.2f} {:>13.0f} {:>6.2f} {:>5.2f}" .format(a6,s6,n6,n6,l6))

# Changing my working directory so I can import the excel file and creating variables to read from the Hotel_Reviews excel sheet

os.chdir("C:\Python 3.6")
openBook = load_workbook('Hotel_Reviews.xlsx')
sheet1 = openBook.active

# Creating a list to count the amount of California data points we have
CAList = []

# Creating a list for california review ratings
reviewRating = []

# Creating a list for the california review sentiments
reviewText = []

# Creating a set to view and count the specific cities within California
cityList = set()

# Creating a separate rating & sentiment list for the specific cities within California
sunnyValeRatingList = []
sunnyValeSentimentList = []

coronadoRatingList = []
coronadoSentimentList = []

grovelandRatingList = []
grovelandSentimentList = []

cypressRatingList = []
cypressSentimentList = []

uplandRatingList = []
uplandSentimentList = []

selmaRatingList = []
selmaSentimentList = []

sanBrunoRatingList = []
sanBrunoSentimentList = []

victorvilleRatingList = []
victorvilleSentimentList = []

livermoreRatingList = []
livermoreSentimentList = []

morroBayRatingList = []
morroBaySentimentList = []

garbervilleRatingList = []
garbervilleSentimentList = []

signalHillRatingList = []
signalHillSentimentList = []

marinaRatingList = []
marinaSentimentList = []

marinaDelReyRatingList = []
marinaDelReySentimentList = []

sanClementeRatingList = []
sanClementeSentimentList = []

napaRatingList = []
napaSentimentList = []

pioneerRatingList = []
pioneerSentimentList = []

tracyRatingList = []
tracySentimentList = []

anaheimRatingList = []
anaheimSentimentList = []

longBeachRatingList = []
longBeachSentimentList = []

barstowRatingList = []
barstowSentimentList = []

sanDiegoRatingList = []
sanDiegoSentimentList = []

buenaParkRatingList = []
buenaParkSentimentList = []

sanJoseRatingList = []
sanJoseSentimentList = []

lakeForestRatingList = []
lakeForestSentimentList = []

hesperiaRatingList = []
hesperiaSentimentList = []

pasadenaRatingList = []
pasadenaSentimentList = []

irvineRatingList = []
irvineSentimentList = []

reddingRatingList = []
reddingSentimentList = []

carlsbadRatingList = []
carlsbadSentimentList = []

studioCityRatingList = []
studioCitySentimentList = []

ranchoMirageRatingList = []
ranchoMirageSentimentList = []

bodegaBayRatingList = []
bodegaBaySentimentList = []

palmSpringsRatingList = []
palmSpringsSentimentList = []

mariposaRatingList = []
mariposaSentimentList = []

losOlivosRatingList = []
losOlivosSentimentList = []

santaBarbaraRatingList = []
santaBarbaraSentimentList = []

stocktonRatingList = []
stocktonSentimentList = []

gardenaRatingList = []
gardenaSentimentList = []

stirlingCityRatingList = []
stirlingCitySentimentList = []

gustineRatingList = []
gustineSentimentList = []

sanFranciscoRatingList = []
sanFranciscoSentimentList = []

willowsRatingList = []
willowsSentimentList = []

gardenGroveRatingList = []
gardenGroveSentimentList = []

joshuaTreeRatingList = []
joshuaTreeSentimentList = []

elCerritoRatingList = []
elCerritoSentimentList = []

eurekaRatingList = []
eurekaSentimentList = []

inglewoodRatingList = []
inglewoodSentimentList = []

sacramentoRatingList = []
sacramentoSentimentList = []

larkspurRatingList = []
larkspurSentimentList = []

coltonRatingList = []
coltonSentimentList = []

rohnertParkRatingList = []
rohnertParkSentimentList = []

# Creating a dictionary to hold all of the cities within california (as the key) and their average sentiment and star rating
caDictRatings = {}
caDictSents = {}

# Adding the values of the sentiments and the ratings to the corresponding dictionaries
caDictRatings["Sunnyvale"] = 3.41
caDictSents["Sunnyvale"] = 0.22
caDictRatings.update({"Coronado" : 4.4, "Groveland" : 0, "Cypress" : 0, "Upland" : 2.11, "Selma" : 0.5, "San Bruno" : 3.96, "Victorville" : 4.04, "Livermore" : 3.83, "Morro Bay" : 3.74, "Garberville" : 3.71, "Signal Hill" : 2.81, "Marina" : 3.24, "Marina Del Rey" : 4.1, "San Clemente" : 4.27, "Napa" : 3.99, "Pioneer" : 0, "Tracy" : 3.67, "Anaheim" : 2.87, "Long Beach" : 3.66, "Barstow" : 3.2, "San Diego" : 3.97, "Buena Park" : 2, "San Jose" : 4.29, "Lake Forest" : 3.19, "Hesperia" : 3.83, "Pasadena" : 3.75, "Irvine" : 4.35, "Redding" : 2.98, "Carlsbad" : 2.98, "Studio City" : 3.21, "Rancho Mirage" : 4.29, "Bodega Bay" : 5, "Palm Springs" : 4.62, "Mariposa" : 2.97, "Los Olivos" : 0, "Santa Barbara" : 0.36, "Stockton" : 0, "Gardena" : 3.45, "Stirling City" : 0, "Gustine" : 0, "San Francisco" : 3.95, "Willows" : 5, "Garden Grove" : 4.33, "Joshua Tree" : 3.75, "El Cerrito" : 0, "Eureka" : 4.23, "Inglewood" : 4, "Sacramento" : 4.95, "Colton" : 3.24, "Rohnert Park" : 3.34})
caDictSents.update({"Coronado" : 0.37, "Groveland" : 0, "Cypress" : 0, "Upland" : 0.1, "Selma" : 0.05, "San Bruno" : 0.33, "Victorville" : 0.28, "Livermore" : 0.31, "Morro Bay" : 0.28, "Garberville" : 0.31, "Signal Hill" : 0.08, "Marina" : 0.08, "Marina Del Rey" : 0.17, "San Clemente" : 0.28, "Napa" : 0.33, "Pioneer" : 0, "Tracy" : 0.26, "Anaheim" : 0.13, "Long Beach" : 0.28, "Barstow" : 0.26, "San Diego" : 0.26, "San Jose" : 0.32, "Lake Forest" : 0.22, "Hesperia" : 0.25, "Pasadena" : 0.29, "Irvine" : 0.35, "Redding" : 0.2, "Carlsbad" : 0.2, "Studio City" : 0.22, "Rancho Mirage" : 0.34, "Bodega Bay" : 0.46, "Palm Springs" : 0.3, "Los Olivos" : 0, "Santa Barbara" : 0.36, "Stockton" : 0, "Gardena" : 0.21, "Stirling City" : 0, "Gustine" : 0, "San Francisco" : 0.31, "Willows" : 0.48, "Garden Grove" : 0.32, "Joshua Tree" : 0.27, "El Cerrito" : 0, "Eureka" : 4.23, "Inglewood" : 0.23, "Sacramento" : 0.36, "Colton" : 0.23, "Rohnert Park" : 0.17})

# Creating a loop to read the whole text file. Within the loop, multiple if-conditions are created to narrow down the specific data we want to analyze
# Provinces are saved into a variable called "province" so we can filter out the ones that are "CA"
# Ratings are saved into a variable called "rating" so we can filter out the ones that are null
# Once we narrowed down California and non-null ratings, we save the California city name in a variable then add it to a set list so print out the unique city names within California
# Another if condition is used for every specific city. That city's corresponding rating & sentiment is added to their list so some descriptive statistics operators can be performed on it later

for i in range(2, 35913):
    province = sheet1.cell(column = 7, row = i).value
    rating = sheet1.cell(column = 10, row = i).value
    
    if rating is not None:
        rating = float(sheet1.cell(column = 10, row = i).value)
    
        if province == "CA":
            
            review = sheet1.cell(column = 11, row = i).value
            textblobText = TextBlob(review)
            sentiment = textblobText.sentiment.polarity
            reviewText.append(sentiment)

            city = sheet1.cell(column = 3, row = i).value
            CAList.append(province)
            reviewRating.append(rating)
            cityList.add(city)

            if city == "Sunnyvale":
                sunnyValeRatingList.append(rating)
                sunnyValeSentimentList.append(sentiment)

            if city == "Coronado":
                coronadoRatingList.append(rating)
                coronadoSentimentList.append(sentiment)

            if city == "Groveland":
                grovelandRatingList.append(rating)
                grovelandSentimentList.append(sentiment)

            if city == "Cypress":
                cypressRatingList.append(rating)
                cypressSentimentList.append(sentiment)

            if city == "Upland":
                uplandRatingList.append(rating)
                uplandSentimentList.append(sentiment)

            if city == "Selma":
                selmaRatingList.append(rating)
                selmaSentimentList.append(sentiment)
                
            if city == "San Bruno":
                sanBrunoRatingList.append(rating)
                sanBrunoSentimentList.append(sentiment)

            if city == "Victorville":
                victorvilleRatingList.append(rating)
                victorvilleSentimentList.append(sentiment)

            if city == "Livermore":
                livermoreRatingList.append(rating)
                livermoreSentimentList.append(sentiment)

            if city == "Morro Bay":
                morroBayRatingList.append(rating)
                morroBaySentimentList.append(sentiment)

            if city == "Garberville":
                garbervilleRatingList.append(rating)
                garbervilleSentimentList.append(sentiment)

            if city == "Signal Hill":
                signalHillRatingList.append(rating)
                signalHillSentimentList.append(sentiment)

            if city == "Marina":
                marinaRatingList.append(rating)
                marinaSentimentList.append(sentiment)

            if city == "Marina Del Rey":
                marinaDelReyRatingList.append(rating)
                marinaDelReySentimentList.append(sentiment)
                
            if city == "San Clemente":
                sanClementeRatingList.append(rating)
                sanClementeSentimentList.append(sentiment)

            if city == "Napa":
                napaRatingList.append(rating)
                napaSentimentList.append(sentiment)

            if city == "Pioneer":
                pioneerRatingList.append(rating)
                pioneerSentimentList.append(sentiment)
                
            if city == "Tracy":
                tracyRatingList.append(rating)
                tracySentimentList.append(sentiment)

            if city == "Anaheim":
                anaheimRatingList.append(rating)
                anaheimSentimentList.append(sentiment)

            if city == "Long Beach":
                longBeachRatingList.append(rating)
                longBeachSentimentList.append(sentiment)

            if city == "Barstow":
                barstowRatingList.append(rating)
                barstowSentimentList.append(sentiment)

            if city == "San Diego":
                sanDiegoRatingList.append(rating)
                sanDiegoSentimentList.append(sentiment)

            if city == "Buena Park":
                buenaParkRatingList.append(rating)
                buenaParkSentimentList.append(sentiment)

            if city == "San Jose":
                sanJoseRatingList.append(rating)
                sanJoseSentimentList.append(sentiment)

            if city == "Lake Forest":
                lakeForestRatingList.append(rating)
                lakeForestSentimentList.append(sentiment)

            if city == "Hesperia":
                hesperiaRatingList.append(rating)
                hesperiaSentimentList.append(sentiment)

            if city == "Pasadena":
                pasadenaRatingList.append(rating)
                pasadenaSentimentList.append(sentiment)

            if city == "Irvine":
                irvineRatingList.append(rating)
                irvineSentimentList.append(sentiment)

            if city == "Redding":
                reddingRatingList.append(rating)
                reddingSentimentList.append(sentiment)

            if city == "Carlsbad":
                carlsbadRatingList.append(rating)
                carlsbadSentimentList.append(sentiment)

            if city == "Studio City":
                studioCityRatingList.append(rating)
                studioCitySentimentList.append(sentiment)

            if city == "Rancho Mirage":
                ranchoMirageRatingList.append(rating)
                ranchoMirageSentimentList.append(sentiment)

            if city == "Bodega Bay":
                bodegaBayRatingList.append(rating)
                bodegaBaySentimentList.append(sentiment)

            if city == "Palm Springs":
                palmSpringsRatingList.append(rating)
                palmSpringsSentimentList.append(sentiment)

            if city == "Mariposa":
                mariposaRatingList.append(rating)
                mariposaSentimentList.append(sentiment)

            if city == "Los Olivos":
                losOlivosRatingList.append(rating)
                losOlivosSentimentList.append(sentiment)

            if city == "Santa Barbara":
                santaBarbaraRatingList.append(rating)
                santaBarbaraSentimentList.append(sentiment)

            if city == "Stockton":
                stocktonRatingList.append(rating)
                stocktonSentimentList.append(sentiment)

            if city == "Gardena":
                gardenaRatingList.append(rating)
                gardenaSentimentList.append(sentiment)

            if city == "Stirling City":
                stirlingCityRatingList.append(rating)
                stirlingCitySentimentList.append(sentiment)

            if city == "Gustine":
                gustineRatingList.append(rating)
                gustineSentimentList.append(sentiment)

            if city == "San Francisco":
                sanFranciscoRatingList.append(rating)
                sanFranciscoSentimentList.append(sentiment)

            if city == "Willows":
                willowsRatingList.append(rating)
                willowsSentimentList.append(sentiment)

            if city == "Garden Grove":
                gardenGroveRatingList.append(rating)
                gardenGroveSentimentList.append(sentiment)

            if city == "Joshua Tree":
                joshuaTreeRatingList.append(rating)
                joshuaTreeSentimentList.append(sentiment)

            if city == "El Cerrito":
                elCerritoRatingList.append(rating)
                elCerritoSentimentList.append(sentiment)

            if city == "Eureka":
                eurekaRatingList.append(rating)
                eurekaSentimentList.append(sentiment)

            if city == "Inglewood":
                inglewoodRatingList.append(rating)
                inglewoodSentimentList.append(sentiment)

            if city == "Sacramento":
                sacramentoRatingList.append(rating)
                sacramentoSentimentList.append(sentiment)

            if city == "Larkspur":
                larkspurRatingList.append(rating)
                larkspurSentimentList.append(sentiment)

            if city == "Colton":
                coltonRatingList.append(rating)
                coltonSentimentList.append(sentiment)

            if city == "Rohnert Park":
                rohnertParkRatingList.append(rating)
                rohnertParkSentimentList.append(sentiment)

# Saving the mean, standard deviation, mode, median, min, and max of the total california review ratings & sentiments to a variable

avgCASent = mean(reviewText)
avgCA = mean(reviewRating)
stDevCASent = stdev(reviewText)
stDevCA = stdev(reviewRating)
modeCASent = mode(reviewText)
modeCA = mode(reviewRating)
medianCASent = median(reviewText)
medianCA = median(reviewRating)
maxCASent = max(reviewText)
maxCA = max(reviewRating)
minCASent = min(reviewText)
minCA = min(reviewRating)
meanSunnyRating = mean(sunnyValeRatingList)
meanSunnySent = mean(sunnyValeSentimentList)

# Printing out the aggregate California statistics values

print("---------------------California ONLY Statistics----------------------")
print("--------------------------Ratings Statistics-------------------------")
print("The Average California Rating Is:",round(avgCA,2))
print("The standard deviation is:",round(stDevCA,2))
print("The mode is:",round(modeCA,2))
print("The median is:",round(medianCA,2))
print("The max is:",round(maxCA,2))
print("The min:",round(minCA,2))
print("---------------------------------------------------------------------")
print("-------------------------Sentiment Statistics------------------------")
print("The average california review is:",round(avgCASent,2))
print("The standard deviation is:",round(stDevCASent,2))
print("The mode is:",round(modeCASent,2))
print("The median is:",round(medianCASent,2))
print("The max is:",round(maxCASent,2))
print("The min:",round(minCASent,2))
print("---------------------------------------------------------------------")

# Printing out the average sentiment and rating values for the specific cities within California

print("---------------Statistics for Cities within California---------------")
meanSunnyRating = mean(sunnyValeRatingList)
meanSunnySent = mean(sunnyValeSentimentList)

print("The average sunnyvale rating is:",round(meanSunnyRating,2))
print("The average review sentiment is:",round(meanSunnySent,2))
print("---------------------------------------------------------------------")

meanCoronadoRating = mean(coronadoRatingList)
meanCoronadoSent = mean(coronadoSentimentList)

print("The average coronado rating is:",round(meanCoronadoRating,2))
print("The average review sentiment is:",round(meanCoronadoSent,2))
print("---------------------------------------------------------------------")

meanGrovelandRating = mean(grovelandRatingList)
meanGrovelandSent = mean(grovelandSentimentList)

print("The average groveland rating is:",round(meanGrovelandRating,2))
print("The average review sentiment is:",round(meanGrovelandSent,2))
print("---------------------------------------------------------------------")

meanCypressRating = mean(cypressRatingList)
meanCypressSent = mean(cypressSentimentList)

print("The average cypress rating is:",round(meanCypressRating,2))
print("The average review sentiment is:",round(meanCypressSent,2))
print("---------------------------------------------------------------------")

meanUplandRating = mean(uplandRatingList)
meanUplandSent = mean(uplandSentimentList)

print("The average upland rating is:",round(meanUplandRating,2))
print("The average review sentiment is:",round(meanUplandSent,2))
print("---------------------------------------------------------------------")

meanSelmaRating = mean(selmaRatingList)
meanSelmaSent = mean(selmaSentimentList)

print("The average selma rating is:",round(meanSelmaRating,2))
print("The average review sentiment is:",round(meanSelmaSent,2))
print("---------------------------------------------------------------------")

meanSanBrunoRating = mean(sanBrunoRatingList)
meanSanBrunoSent = mean(sanBrunoSentimentList)

print("The average san bruno rating is:",round(meanSanBrunoRating,2))
print("The average review sentiment is:",round(meanSanBrunoSent,2))
print("---------------------------------------------------------------------")

meanVictorvilleRating = mean(victorvilleRatingList)
meanVictorvilleSent = mean(victorvilleSentimentList)

print("The average victorville rating is:",round(meanVictorvilleRating,2))
print("The average review sentiment is:",round(meanVictorvilleSent,2))
print("---------------------------------------------------------------------")

meanLivermoreRating = mean(livermoreRatingList)
meanLivermoreSent = mean(livermoreSentimentList)

print("The average livermore rating is:",round(meanLivermoreRating,2))
print("The average review sentiment is:",round(meanLivermoreSent,2))
print("---------------------------------------------------------------------")

meanMorroBayRating = mean(morroBayRatingList)
meanMorroBaySent = mean(morroBaySentimentList)

print("The average morro bay rating is:",round(meanMorroBayRating,2))
print("The average review sentiment is:",round(meanMorroBaySent,2))
print("---------------------------------------------------------------------")

meanGarbervilleRating = mean(garbervilleRatingList)
meanGarbervilleSent = mean(garbervilleSentimentList)

print("The average garberville rating is:",round(meanGarbervilleRating,2))
print("The average review sentiment is:",round(meanGarbervilleSent,2))
print("---------------------------------------------------------------------")

meanSignalHillRating = mean(signalHillRatingList)
meanSignalHillSent = mean(signalHillSentimentList)

print("The average signal hill rating is:",round(meanSignalHillRating,2))
print("The average review sentiment is:",round(meanSignalHillSent,2))
print("---------------------------------------------------------------------")

meanMarinaRating = mean(marinaRatingList)
meanMarinaSent = mean(marinaSentimentList)

print("The average marina rating is:",round(meanMarinaRating,2))
print("The average review sentiment is:",round(meanMarinaSent,2))
print("---------------------------------------------------------------------")

meanMarinaDelReyRating = mean(marinaDelReyRatingList)
meanMarinaDelReySent = mean(marinaDelReySentimentList)

print("The average marina del rey rating is:",round(meanMarinaDelReyRating,2))
print("The average review sentiment is:",round(meanMarinaDelReySent,2))
print("---------------------------------------------------------------------")

meanSanClementeRating = mean(sanClementeRatingList)
meanSanClementeSent = mean(sanClementeSentimentList)

print("The average san clemente rating is:",round(meanSanClementeRating,2))
print("The average review sentiment is:",round(meanSanClementeSent,2))
print("---------------------------------------------------------------------")

meanNapaRating = mean(napaRatingList)
meanNapaSent = mean(napaSentimentList)

print("The average napa rating is:",round(meanNapaRating,2))
print("The average review sentiment is:",round(meanNapaSent,2))
print("---------------------------------------------------------------------")

meanPioneerRating = mean(pioneerRatingList)
meanPioneerSent = mean(pioneerSentimentList)

print("The average pioneer rating is:",round(meanPioneerRating,2))
print("The average review sentiment is:",round(meanPioneerSent,2))
print("---------------------------------------------------------------------")

meanTracyRating = mean(tracyRatingList)
meanTracySent = mean(tracySentimentList)

print("The average tracy rating is:",round(meanTracyRating,2))
print("The average review sentiment is:",round(meanTracySent,2))
print("---------------------------------------------------------------------")

meanAnaheimRating = mean(anaheimRatingList)
meanAnaheimSent = mean(anaheimSentimentList)

print("The average anaheim rating is:",round(meanAnaheimRating,2))
print("The average review sentiment is:",round(meanAnaheimSent,2))
print("---------------------------------------------------------------------")

meanLongBeachRating = mean(longBeachRatingList)
meanLongBeachSent = mean(longBeachSentimentList)

print("The average long beach rating is:",round(meanLongBeachRating,2))
print("The average review sentiment is:",round(meanLongBeachSent,2))
print("---------------------------------------------------------------------")

meanBarstowRating = mean(barstowRatingList)
meanBarstowSent = mean(barstowSentimentList)

print("The average barstow rating is:",round(meanBarstowRating,2))
print("The average review sentiment is:",round(meanBarstowSent,2))
print("---------------------------------------------------------------------")

meanSanDiegoRating = mean(sanDiegoRatingList)
meanSanDiegoSent = mean(sanDiegoSentimentList)

print("The average san diego rating is:",round(meanSanDiegoRating,2))
print("The average review sentiment is:",round(meanSanDiegoSent,2))
print("---------------------------------------------------------------------")

meanBuenaParkRating = mean(buenaParkRatingList)
meanBuenaParkSent = mean(buenaParkSentimentList)

print("The average buena park rating is:",round(meanBuenaParkRating,2))
print("The average review sentiment is:",round(meanBuenaParkSent,2))
print("---------------------------------------------------------------------")

meanSanJoseRating = mean(sanJoseRatingList)
meanSanJoseSent = mean(sanJoseSentimentList)

print("The average san jose rating is:",round(meanSanJoseRating,2))
print("The average review sentiment is:",round(meanSanJoseSent,2))
print("---------------------------------------------------------------------")

meanLakeForestRating = mean(lakeForestRatingList)
meanLakeForestSent = mean(lakeForestSentimentList)

print("The average lake forest rating is:",round(meanLakeForestRating,2))
print("The average review sentiment is:",round(meanLakeForestSent,2))
print("---------------------------------------------------------------------")

meanHesperiaRating = mean(hesperiaRatingList)
meanHesperiaSent = mean(hesperiaSentimentList)

print("The average hesperia rating is:",round(meanHesperiaRating,2))
print("The average review sentiment is:",round(meanHesperiaSent,2))
print("---------------------------------------------------------------------")

meanPasadenaRating = mean(pasadenaRatingList)
meanPasadenaSent = mean(pasadenaSentimentList)

print("The average pasadena rating is:",round(meanPasadenaRating,2))
print("The average review sentiment is:",round(meanPasadenaSent,2))
print("---------------------------------------------------------------------")

irvineRatingListAvg = mean(irvineRatingList)
irvineSentimentListAvg = mean(irvineSentimentList)

print("The average Irvine rating is: ", round(irvineRatingListAvg, 2))
print("The average review sentiment is: ", round(irvineSentimentListAvg, 2))
print("-------------------------------------------------------------")

reddingRatingListAvg = mean(carlsbadRatingList)
reddingSentimentListAvg = mean(carlsbadSentimentList)

print("The average Redding rating is: ", round(reddingRatingListAvg, 2))
print("The average review sentiment is: ", round(reddingSentimentListAvg, 2))
print("-------------------------------------------------------------")

carlsbadRatingListAvg = mean(carlsbadRatingList)
carlsbadSentimentListAvg = mean(carlsbadSentimentList)

print("The average Carlsbad rating is: ", round(carlsbadRatingListAvg, 2))
print("The average review sentiment is: ", round(carlsbadSentimentListAvg, 2))
print("-------------------------------------------------------------")

studioCityRatingListAvg = mean(studioCityRatingList)
studioCitySentimentListAvg = mean(studioCitySentimentList)

print("The average Studio City rating is: ", round(studioCityRatingListAvg, 2))
print("The average review sentiment is: ", round(studioCitySentimentListAvg, 2))
print("-------------------------------------------------------------")

ranchoMirageRatingListAvg = mean(ranchoMirageRatingList)
ranchoMirageSentimentListAvg = mean(ranchoMirageSentimentList)

print("The average Rancho Mirage rating is: ", round(ranchoMirageRatingListAvg, 2))
print("The average review sentiment is: ", round(ranchoMirageSentimentListAvg, 2))
print("-------------------------------------------------------------")

bodegaBayRatingListAvg = mean(bodegaBayRatingList)
bodegaBaySentimentListAvg = mean(bodegaBaySentimentList)

print("The average Bodega Bay rating is: ", round(bodegaBayRatingListAvg, 2))
print("The average review sentiment is: ", round(bodegaBaySentimentListAvg, 2))
print("-------------------------------------------------------------")

palmSpringsRatingListAvg = mean(palmSpringsRatingList)
palmSpringsSentimentListAvg = mean(palmSpringsSentimentList)

print("The average Palm Springs rating is: ", round(palmSpringsRatingListAvg, 2))
print("The average review sentiment is: ", round(palmSpringsSentimentListAvg, 2))
print("-------------------------------------------------------------")

mariposaRatingListAvg = mean(mariposaRatingList)
mariposaSentimentListAvg = mean(mariposaSentimentList)

print("The average Mariposa rating is: ", round(mariposaRatingListAvg, 2))
print("The average review sentiment is: ", round(mariposaSentimentListAvg, 2))
print("-------------------------------------------------------------")

losOlivosRatingListAvg = mean(losOlivosRatingList)
losOlivosSentimentListAvg = mean(losOlivosSentimentList)

print("The average Los Olivos rating is: ", round(losOlivosRatingListAvg, 2))
print("The average review sentiment is: ", round(losOlivosSentimentListAvg, 2))
print("-------------------------------------------------------------")

santaBarbaraRatingListAvg = mean(santaBarbaraSentimentList)
santaBarbaraSentimentListAvg = mean(santaBarbaraSentimentList)

print("The average Santa Barbara rating is: ", round(santaBarbaraRatingListAvg, 2))
print("The average review sentiment is: ", round(santaBarbaraSentimentListAvg, 2))
print("-------------------------------------------------------------")

stocktonRatingListAvg = mean(stocktonRatingList)
stocktonSentimentListAvg = mean(stocktonSentimentList)

print("The average Stockton rating is: ", round(stocktonRatingListAvg, 2))
print("Average review sentiment is: ", round(stocktonSentimentListAvg, 2))
print("-------------------------------------------------------------")

gardenaRatingListAvg = mean(gardenaRatingList)
gardenaSentimentListAvg = mean(gardenaSentimentList)

print("The average Gardena rating is: ", round(gardenaRatingListAvg, 2))
print("The average review sentiment is: ", round(gardenaSentimentListAvg, 2))
print("-------------------------------------------------------------")

stirlingCityRatingListAvg = mean(stirlingCityRatingList)
stirlingCitySentimentListAvg = mean(stirlingCitySentimentList)

print("The average Stirling City rating is: ", round(stirlingCityRatingListAvg, 2))
print("The average review sentiment is: ", round(stirlingCitySentimentListAvg, 2))
print("-------------------------------------------------------------")

gustineRatingListAvg = mean(gustineRatingList)
gustineSentimentListAvg = mean(gustineSentimentList)

print("The average Gustine rating is: ", round(gustineRatingListAvg, 2))
print("Average review sentiment is: ", round(gustineSentimentListAvg, 2))
print("-------------------------------------------------------------")

sanFranciscoRatingListAvg = mean(sanFranciscoRatingList)
sanFranciscoSentimentListAvg = mean(sanFranciscoSentimentList)

print("The average San Franciso rating is: ", round(sanFranciscoRatingListAvg, 2))
print("The average review sentiment is: ", round(sanFranciscoSentimentListAvg, 2))
print("-------------------------------------------------------------")

willowsRatingListAvg = mean(willowsRatingList)
willowsSentimentListAvg = mean(willowsSentimentList)

print("The average Willows rating is: ", round(willowsRatingListAvg, 2))
print("The average review sentiment is: ", round(willowsSentimentListAvg, 2))
print("-------------------------------------------------------------")

gardenGroveRatingListAvg = mean(gardenGroveRatingList)
gardenGroveSentimentListAvg = mean(gardenGroveSentimentList)

print("The average Garden Grove rating is: ", round(gardenGroveRatingListAvg, 2))
print("The average review sentiment is: ", round(gardenGroveSentimentListAvg, 2))
print("-------------------------------------------------------------")

joshuaTreeRatingListAvg = mean(joshuaTreeRatingList)
joshuaTreeSentimentListAvg = mean(joshuaTreeSentimentList)

print("The average Joshua Tree rating is: ", round(joshuaTreeRatingListAvg, 2))
print("The average review sentiment is: ", round(joshuaTreeSentimentListAvg, 2))
print("-------------------------------------------------------------")

elCerritoRatingListAvg = mean(elCerritoRatingList)
elCerritoSentimentListAvg = mean(elCerritoSentimentList)

print("The average El Cerrito rating is: ", round(elCerritoRatingListAvg, 2))
print("The average review sentiment is: ", round(elCerritoSentimentListAvg, 2))
print("-------------------------------------------------------------")

eurekaRatingListAvg = mean(eurekaRatingList)
eurekaSentimentListAvg = mean(eurekaSentimentList)

print("The average Eureka rating is: ", round(eurekaRatingListAvg, 2))
print("The verage review sentiment is: ", round(eurekaSentimentListAvg, 2))
print("-------------------------------------------------------------")

inglewoodRatingListAvg = mean(inglewoodRatingList)
inglewoodSentimentListAvg = mean(inglewoodSentimentList)

print("The average Inglewood rating is: ", round(inglewoodRatingListAvg, 2))
print("The average review sentiment is: ", round(inglewoodSentimentListAvg, 2))
print("-------------------------------------------------------------")

sacramentoRatingListAvg = mean(sacramentoRatingList)
sacramentoSentimentListAvg = mean(sacramentoSentimentList)

print("The average Sacramento rating is: ", round(sacramentoRatingListAvg, 2))
print("The average review sentiment is: ", round(sacramentoSentimentListAvg, 2))
print("-------------------------------------------------------------")

coltonRatingListAvg = mean(coltonRatingList)
coltonSentimentListAvg = mean(coltonSentimentList)

print("The average Colton rating is: ", round(coltonRatingListAvg, 2))
print("The average review sentiment is: ", round(coltonSentimentListAvg, 2))
print("-------------------------------------------------------------")


rohnertParkRatingListAvg = mean(rohnertParkRatingList)
rohnertParkSentimentListAvg = mean(rohnertParkSentimentList)

print("The average Rohnert Park rating is: ", round(rohnertParkRatingListAvg, 2))
print("The average review sentiment is: ", round(rohnertParkSentimentListAvg, 2))
print("-------------------------------------------------------------")
print()

# Printing out the length of our CAList list to find the amount of California data points we have
# Printing out the length of our cityList list to find the amount of "cities within California" datapoints we have
# Lastly, printing out the cityList ITSELF to find what specific cities we have within California. This is done so the user knows what cities they can look up in the bottom function

print("---------------------Sample Size Statistics---------------------")
print("The amount of California Datapoints is:",len(CAList))
print("The amount of Cities in California is:",len(cityList))
print(" ")
print("These are the cities available, to choose from, in California:",cityList)
print("-----------------------------------------------------------------")

# Using the nlargest and nsmallest function, from the imported heapq library, to find the 3 highest and 3 lowest average review ratings & sentiments for the cities then assigning them to a variable and printing it

threeLargestReviews = nlargest(3, caDictRatings, key = caDictRatings.get)
threeLowestReviews = nsmallest(3, caDictRatings, key = caDictRatings.get)
print("The three highest reviewed cities, according to average ratings, are: ",threeLargestReviews)
print("The three lowest reviewed cities, according to average ratings, are: ",threeLowestReviews)
print()
threeLargestSents = nlargest(3, caDictRatings, key = caDictRatings.get)
threeLowestSents = nsmallest(3, caDictRatings, key = caDictRatings.get)
print("The three highest reviewed cities, according to sentiments, are: ",threeLargestSents)
print("The three lowest reviewed cities, according to sentiments, are: ", threeLowestSents)
print()

# Using our user defined function called "six_table_creatorReviews" and "six_table_creatorSents" to create a table for the descriptive statistics of our top 3 and bottom 3 average ratings & sentiments cities

six_table_creatorReviews(threeLargestSents, threeLowestSents, mean(bodegaBayRatingList), stdev(bodegaBayRatingList), len(bodegaBayRatingList), max(bodegaBayRatingList), min(bodegaBayRatingList), mean(willowsRatingList), 0, len(willowsRatingList), max(willowsRatingList), min(willowsRatingList), mean(sacramentoRatingList), stdev(sacramentoRatingList), len(sacramentoRatingList), max(sacramentoRatingList), min(sacramentoRatingList), mean(grovelandRatingList), 0, len(grovelandRatingList), max(grovelandRatingList), min(grovelandRatingList), mean(cypressRatingList), 0, len(cypressRatingList), max(cypressRatingList), min(cypressRatingList), mean(pioneerRatingList), 0, len(pioneerRatingList), max(pioneerRatingList), min(pioneerRatingList))

print()

six_table_creatorSents(threeLargestSents, threeLowestSents, mean(bodegaBaySentimentList), stdev(bodegaBaySentimentList), len(bodegaBaySentimentList), max(bodegaBaySentimentList), min(bodegaBaySentimentList), mean(willowsSentimentList), 0, len(willowsSentimentList), max(willowsSentimentList), min(willowsSentimentList), mean(sacramentoSentimentList), stdev(sacramentoSentimentList), len(sacramentoSentimentList), max(sacramentoSentimentList), min(sacramentoSentimentList), mean(grovelandSentimentList), 0, len(grovelandSentimentList), max(grovelandSentimentList), min(grovelandSentimentList), mean(cypressSentimentList), 0, len(cypressSentimentList), max(cypressSentimentList), min(cypressSentimentList), mean(pioneerSentimentList), 0, len(pioneerSentimentList), max(pioneerSentimentList), min(pioneerSentimentList))


print()

# Creating a separate book and sheet object to write into an excel file

book = Workbook()
sheet = book.active

sheet['A1'] = "Bodega Bay Ratings"
sheet['B1'] = "Bodega Bay Sentiments"
sheet['D1'] = "Sacramento Ratings"
sheet['E1'] = "Sacramento Sentiments"
sheet['G1'] = "Bin Range Ratings" 
sheet['H1'] = "Bin Range Sentiments"

# Creating separate loops to add values of the bodega bay and sacramento ratings/sentiments + bin ranges to create a histogram on excel

row = 2
for i, value in enumerate(bodegaBayRatingList):
    sheet.cell(column = 1, row = row + i, value = value)

for i, value in enumerate(bodegaBaySentimentList):
    sheet.cell(column = 2, row = row + i, value = value)

for i, value in enumerate(sacramentoRatingList):
    sheet.cell(column = 4, row = row + i, value = value)

for i, value in enumerate(sacramentoSentimentList):
    sheet.cell(column = 5, row = row + i, value = value)

binRangeRatings = [1,2,3,4,5]
for i, value in enumerate(binRangeRatings):
    sheet.cell(column = 7, row = row + i, value = value)

binRangeSentiments = [-0.5,-0.3,-0.1,.1,.3,.5]
for i, value in enumerate(binRangeSentiments):
    sheet.cell(column = 8, row = row + i, value = value)

book.save('Project.xlsx')


# Creating a lookup function to allow customers to look up the statistics of a city within California
# Create loop and plan for user error


while True:
    try:
        userInput = int(input("If you'd like to see the rating and sentiment for a specific city, enter 1. If not, enter 2 to exit: "))
    except ValueError:
        print("Sorry, please enter the number 1 or 2.")
        continue
    if userInput == 1:
        userCity = input("Type in the city you wish to learn about (Please use correct spacing and capitalization): ")
        if userCity not in cityList:
            print("You entered the wrong city name! Try again!")
            continue
        elif userCity in cityList:
            print()
            print("The star rating for", userCity, "is: ", caDictRatings[userCity])
            print("The sentiment for", userCity, "is: ", caDictSents[userCity])
            print()
            continue
    elif userInput == 2:
        print("Godspeed!!")
    else:
        print("You didn't enter the number 1 or 2. Please try again.")
        continue
    break






